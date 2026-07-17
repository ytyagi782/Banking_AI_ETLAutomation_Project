"""
excel_report.py
---------------
Builds TWO of the three reports:

  1. Summary report   -> high level pass / fail overview
  2. Detailed report  -> every failure, with coloured highlighting:
        yellow  = column value differs between source and target
        red     = record missing on one side
        orange  = duplicate record

(The third report - the management HTML - is in html_report.py.)
"""

import os
import getpass
from datetime import datetime, date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

from utilities import result_store, config_loader
from utilities.config_loader import get_settings

_settings = get_settings()
_colors = _settings["reporting"]["colors"]
# branding block shared with the HTML report (may be empty)
_BRAND = (_settings.get("reporting", {}) or {}).get("branding", {}) or {}

# reusable styles ----------------------------------------------------------
FILL_DIFF = PatternFill("solid", fgColor=_colors["diff"])       # yellow
FILL_MISSING = PatternFill("solid", fgColor=_colors["missing"])  # red
FILL_DUP = PatternFill("solid", fgColor=_colors["duplicate"])    # orange
FILL_HEADER = PatternFill("solid", fgColor="305496")             # blue header
FILL_PASS = PatternFill("solid", fgColor="C6EFCE")               # green
FILL_FAIL = PatternFill("solid", fgColor="FFC7CE")               # light red
FILL_SKIP = PatternFill("solid", fgColor="D9D9D9")               # grey (skipped)

FONT_HEADER = Font(bold=True, color="FFFFFF")
FONT_TITLE = Font(bold=True, size=14)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---- branding banner styles (mirror the HTML management report header) ----
FILL_BRAND = PatternFill("solid", fgColor="102A43")              # dark navy
FONT_BRAND_TITLE = Font(bold=True, size=16, color="FFFFFF")      # white title
FONT_BRAND_COMPANY = Font(bold=True, size=11, color="FFD27D")    # gold company
FONT_BRAND_META = Font(size=9, color="C3D0E8")                   # light-blue meta
FONT_BRAND_AUTHOR = Font(bold=True, size=10, color="FFFFFF")     # author name
FONT_BRAND_ROLE = Font(size=9, color="C3D0E8")                   # author role

# rows occupied by the banner; content is written below it
_BRAND_ROWS = 5
_BRAND_SPAN = 8             # default number of columns the banner spans
DEFAULT_TITLE = "Banking ETL Automation - Test Execution Report"


def _safe(value):
    """Convert a value to something openpyxl can write."""
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return str(value)
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


def _sheet_name(text):
    """Make a safe, unique-ish sheet name (Excel max 31 chars)."""
    bad = '[]:*?/\\'
    clean = "".join("_" if c in bad else c for c in text)
    return clean[:31]


def _write_header(ws, headers, row=1):
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


def _autofit(ws, max_width=50, skip_rows=_BRAND_ROWS):
    """
    Size columns to their content. Merged cells (used by the banner) and the
    banner rows themselves are ignored so the long title text does not blow up
    a data column's width.
    """
    widths = {}
    for row in ws.iter_rows(min_row=skip_rows + 1):
        for cell in row:
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            letter = cell.column_letter
            widths[letter] = max(widths.get(letter, 0), len(str(cell.value)))
    for letter, length in widths.items():
        ws.column_dimensions[letter].width = min(max_width, max(12, length + 2))


# ==========================================================================
# BRANDING BANNER  (same look as the management HTML report header)
# ==========================================================================
def _current_user():
    """Best-effort system username for the report header."""
    try:
        return getpass.getuser()
    except Exception:
        return "unknown"


def _load_image(path, height):
    """
    Return an openpyxl Image for `path`, scaled to `height` px (aspect ratio
    kept), or None if the path is blank / missing / cannot be loaded. Mirrors
    the HTML report's _data_uri: a missing image simply omits the element.
    """
    if not path:
        return None
    abs_path = path if os.path.isabs(path) else config_loader.abs_path(path)
    if not os.path.exists(abs_path):
        return None
    try:
        img = XLImage(abs_path)
    except Exception:
        return None
    if img.height:
        ratio = height / float(img.height)
        img.width = int(img.width * ratio)
        img.height = height
    return img


def _merged(ws, r1, c1, r2, c2, value, font, align="left"):
    """Merge a cell range, write the top-left cell and style it."""
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    cell.font = font
    cell.alignment = Alignment(horizontal=align, vertical="center")
    return cell


def _write_branding_header(ws, title=DEFAULT_TITLE, span=_BRAND_SPAN):
    """
    Paint a branded banner across rows 1.._BRAND_ROWS:
      logo + company (left) | title + generated meta (centre) | author (right)
    Returns the first free row below the banner (callers offset content there).
    """
    span = max(span, _BRAND_SPAN)

    # banner background + row heights
    for r in range(1, _BRAND_ROWS + 1):
        ws.row_dimensions[r].height = 20
        for c in range(1, span + 1):
            ws.cell(row=r, column=c).fill = FILL_BRAND

    # company logo (left)
    logo = _load_image(_BRAND.get("company_logo"), height=60)
    if logo:
        ws.add_image(logo, "A1")

    # centre text block: title / company / generated on / generated by
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    user = _current_user()
    company_name = _BRAND.get("company_name") or ""
    text_end = max(5, span - 3)                 # leave room for the author block
    _merged(ws, 1, 3, 2, text_end, title, FONT_BRAND_TITLE)
    if company_name:
        _merged(ws, 3, 3, 3, text_end, company_name, FONT_BRAND_COMPANY)
    _merged(ws, 4, 3, 4, text_end, f"Generated on : {now}", FONT_BRAND_META)
    _merged(ws, 5, 3, 5, text_end, f"Generated by : {user}", FONT_BRAND_META)

    # author block (right): name + role text, photo at the far edge
    author_name = _BRAND.get("author_name") or user
    author_role = _BRAND.get("author_role") or "Report Author"
    auth_start = span - 2
    _merged(ws, 2, auth_start, 2, span - 1, author_name, FONT_BRAND_AUTHOR, "right")
    _merged(ws, 3, auth_start, 3, span - 1, author_role, FONT_BRAND_ROLE, "right")
    author_img = _load_image(_BRAND.get("author_image"), height=60)
    if author_img:
        ws.add_image(author_img, f"{get_column_letter(span)}1")

    return _BRAND_ROWS + 2                       # one blank spacer row below


# ==========================================================================
# 1. SUMMARY REPORT
# ==========================================================================
def build_summary_report(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # branded banner; everything else is shifted down by `off`
    off = _write_branding_header(ws) - 1

    counts = result_store.summary_counts()
    ws.cell(row=1 + off, column=1,
            value="Banking ETL Automation - Summary Report").font = FONT_TITLE
    ws.cell(row=2 + off, column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # overall totals
    ws.cell(row=4 + off, column=1, value="Total")
    ws.cell(row=4 + off, column=2, value=counts["total"])
    ws.cell(row=5 + off, column=1, value="Passed")
    ws.cell(row=5 + off, column=2, value=counts["passed"]).fill = FILL_PASS
    ws.cell(row=6 + off, column=1, value="Failed")
    fcell = ws.cell(row=6 + off, column=2, value=counts["failed"])
    if counts["failed"]:
        fcell.fill = FILL_FAIL
    ws.cell(row=7 + off, column=1, value="Skipped")
    scell = ws.cell(row=7 + off, column=2, value=counts["skipped"])
    if counts["skipped"]:
        scell.fill = FILL_SKIP
    ws.cell(row=8 + off, column=1, value="Pass rate %")
    ws.cell(row=8 + off, column=2, value=counts["pass_rate"])
    for r in range(4 + off, 9 + off):
        ws.cell(row=r, column=1).font = Font(bold=True)

    # by layer
    ws.cell(row=10 + off, column=1, value="Results by layer").font = Font(bold=True)
    _write_header(ws, ["Layer", "Passed", "Failed", "Skipped"], row=11 + off)
    r = 12 + off
    for layer, g in result_store.summary_by_layer().items():
        ws.cell(row=r, column=1, value=layer)
        ws.cell(row=r, column=2, value=g["passed"])
        fcell = ws.cell(row=r, column=3, value=g["failed"])
        if g["failed"]:
            fcell.fill = FILL_FAIL
        scell = ws.cell(row=r, column=4, value=g["skipped"])
        if g["skipped"]:
            scell.fill = FILL_SKIP
        r += 1

    # full result list
    start = r + 2
    ws.cell(row=start - 1, column=1, value="All validations").font = Font(bold=True)
    _write_header(ws, ["Layer", "Table", "Validation", "Category",
                       "Status", "Message"], row=start)
    row = start + 1
    status_fill = {"PASS": FILL_PASS, "FAIL": FILL_FAIL, "SKIP": FILL_SKIP}
    # executed tests first (PASS/FAIL), then the skipped ones
    for res in result_store.get_executed() + result_store.get_skipped():
        ws.cell(row=row, column=1, value=res["layer"])
        ws.cell(row=row, column=2, value=res["table"])
        ws.cell(row=row, column=3, value=res["validation"])
        ws.cell(row=row, column=4, value=res["category"])
        scell = ws.cell(row=row, column=5, value=res["status"])
        scell.fill = status_fill.get(res["status"], FILL_FAIL)
        ws.cell(row=row, column=6, value=res["message"])
        row += 1

    _autofit(ws)
    wb.save(path)
    return path


# ==========================================================================
# 2. DETAILED REPORT  (coloured)
# ==========================================================================
def _add_failures_sheet(wb):
    ws = wb.active
    ws.title = "Failures"
    off = _write_branding_header(ws) - 1
    ws.cell(row=1 + off, column=1, value="Failed Validations").font = FONT_TITLE
    _write_header(ws, ["Layer", "Table", "Validation", "Category", "Message"],
                  row=3 + off)
    failures = result_store.get_failures()
    if not failures:
        cell = ws.cell(row=5 + off, column=1,
                       value="No failures - all validations passed.")
        cell.fill = FILL_PASS
        return
    row = 4 + off
    for f in failures:
        ws.cell(row=row, column=1, value=f["layer"])
        ws.cell(row=row, column=2, value=f["table"])
        ws.cell(row=row, column=3, value=f["validation"])
        ws.cell(row=row, column=4, value=f["category"])
        ws.cell(row=row, column=5, value=f["message"]).fill = FILL_FAIL
        row += 1
    _autofit(ws)


# friendly layer labels shown in the "Layer" column
_LAYER_LABEL = {
    "SourceToPreStaging": "Source -> PreStaging",
    "PreStagingToStaging": "PreStaging -> Staging",
    "StagingToDWH": "Staging -> DWH",
}

# the leading (extra) columns added in front of the real table columns
_META_HEADERS = ["Layer", "FailureType", "RowDestination"]

# which highlight colour goes with which FailureType
_TYPE_FILL = {
    "DF": FILL_DIFF,                    # yellow
    "Missing in Source": FILL_MISSING,  # red
    "Missing in Target": FILL_MISSING,  # red
    "Duplicate in Source": FILL_DUP,    # orange
    "Duplicate in Target": FILL_DUP,    # orange
}


def _has_issues(detail):
    return bool(detail["diff_pairs"] or detail["missing_in_target"]
                or detail["missing_in_source"] or detail["duplicates"])


def _display_columns(details):
    """Ordered union of the real table columns across all of a table's layers."""
    cols = []
    for d in details:
        for c in d.get("columns", []):
            if c not in cols:
                cols.append(c)
    return cols


def _failure_records(detail):
    """
    Turn one comparison into failure records. Each record is:
      { layer, ftype, dest, row (dict of column->value), highlight (set of cols) }
    so the report can lay the real columns out in their own cells.
    """
    layer = _LAYER_LABEL.get(detail["layer"], detail["layer"])
    recs = []

    # DF = value differs -> show the FULL source row and the FULL target row,
    # highlighting only the columns that actually differ.
    for pair in detail["diff_pairs"]:
        hl = set(pair["diff_columns"])
        recs.append({"layer": layer, "ftype": "DF", "dest": "Source",
                     "row": pair["source_row"], "highlight": hl})
        recs.append({"layer": layer, "ftype": "DF", "dest": "Target",
                     "row": pair["target_row"], "highlight": hl})

    # record exists in source but NOT in target (whole row highlighted)
    for row in detail["missing_in_target"]:
        recs.append({"layer": layer, "ftype": "Missing in Target", "dest": "Source",
                     "row": row, "highlight": set(row.keys())})

    # record exists in target but NOT in source
    for row in detail["missing_in_source"]:
        recs.append({"layer": layer, "ftype": "Missing in Source", "dest": "Target",
                     "row": row, "highlight": set(row.keys())})

    # duplicate key (the stored _side says which table it was on)
    for row in detail["duplicates"]:
        side = row.get("_side", "source")
        dest = "Source" if side == "source" else "Target"
        clean = {c: v for c, v in row.items() if not c.startswith("_")}
        recs.append({"layer": layer, "ftype": f"Duplicate in {dest}", "dest": dest,
                     "row": clean, "highlight": set(clean.keys())})

    return recs


def _add_table_sheet(wb, table, details, max_rows):
    """
    ONE sheet per table, laid out like a normal Excel table:
    Layer | FailureType | RowDestination | <each real table column ...>
    Cells are coloured: yellow=differing value, red=missing row, orange=duplicate.
    """
    records = []
    for detail in details:                      # details are in layer order
        if _has_issues(detail):
            records.extend(_failure_records(detail))
    if not records:
        return  # this table passed everywhere - no sheet needed
    records = records[:max_rows]

    display_cols = _display_columns(details)
    n_meta = len(_META_HEADERS)

    ws = wb.create_sheet(_sheet_name(table))
    # banner spans the full table width; content shifts down by `off`
    off = _write_branding_header(ws, span=n_meta + len(display_cols)) - 1
    ws.cell(row=1 + off, column=1,
            value=f"Detailed failures - {table}").font = FONT_TITLE

    # legend explaining the colours / failure types
    ws.cell(row=2 + off, column=1, value="Legend:")
    ws.cell(row=2 + off, column=2, value="DF = value differs").fill = FILL_DIFF
    ws.cell(row=2 + off, column=3, value="Missing record").fill = FILL_MISSING
    ws.cell(row=2 + off, column=4, value="Duplicate record").fill = FILL_DUP

    # header row = extra columns + the real table columns
    _write_header(ws, _META_HEADERS + display_cols, row=4 + off)

    # data rows
    r = 5 + off
    for rec in records:
        fill = _TYPE_FILL.get(rec["ftype"])
        ws.cell(row=r, column=1, value=rec["layer"])
        ftype_cell = ws.cell(row=r, column=2, value=rec["ftype"])
        if fill:
            ftype_cell.fill = fill
        ws.cell(row=r, column=3, value=rec["dest"])
        for i, col in enumerate(display_cols, start=n_meta + 1):
            cell = ws.cell(row=r, column=i, value=_safe(rec["row"].get(col)))
            if fill and col in rec["highlight"]:
                cell.fill = fill
        r += 1

    _autofit(ws)


def build_detailed_report(path):
    wb = Workbook()
    _add_failures_sheet(wb)
    max_rows = _settings["reporting"].get("max_detail_rows", 500)

    # group all comparisons by table name  ->  exactly ONE sheet per table
    by_table = {}
    for detail in result_store.get_comparisons().values():
        by_table.setdefault(detail["table"], []).append(detail)

    for table, details in by_table.items():
        _add_table_sheet(wb, table, details, max_rows)

    wb.save(path)
    return path
