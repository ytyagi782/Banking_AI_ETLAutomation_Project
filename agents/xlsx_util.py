"""
agents/xlsx_util.py
===================
Tiny shared helper for writing clean, consistently-styled Excel sheets from the
agents (mapping document, test cases, defects). Keeps the openpyxl boilerplate
in one place so every agent's workbook looks the same:

  * bold white header on a dark-blue band, frozen at the top
  * thin borders, wrapped text, sensible auto-fitted column widths
  * an optional "group band" row above the header (used by the mapping doc to
    label the SOURCE / TRANSFORM / TARGET column groups)

Colours match the framework's palette family (see config/settings.yaml).
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")   # dark blue
GROUP_FILLS = {
    "source": PatternFill("solid", fgColor="2E75B6"),   # medium blue
    "transform": PatternFill("solid", fgColor="C55A11"),  # orange
    "target": PatternFill("solid", fgColor="548235"),   # green
}
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
GROUP_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP_TOP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

MAX_WIDTH = 55
MIN_WIDTH = 10


def write_sheet(ws, headers, rows, group_bands=None, widths=None):
    """
    Write a styled table to worksheet `ws`.

    headers      : list[str]                - the column titles
    rows         : list[list]               - the data (each inner list a row)
    group_bands  : optional list[(label, span, key)] drawn ABOVE the header,
                   e.g. [("SOURCE", 7, "source"), ("TRANSFORM", 1, "transform"),
                         ("TARGET", 7, "target")]. key selects a GROUP_FILLS colour.
    widths       : optional list[int] fixed column widths; else auto-fit.
    """
    start_row = 1

    # ---- optional group band row ----
    if group_bands:
        col = 1
        for label, span, key in group_bands:
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = GROUP_FONT
            cell.fill = GROUP_FILLS.get(key, HEADER_FILL)
            cell.alignment = CENTER
            cell.border = BORDER
            if span > 1:
                ws.merge_cells(start_row=1, start_column=col,
                               end_row=1, end_column=col + span - 1)
                # style the merged-over cells too (for borders)
                for c in range(col + 1, col + span):
                    mc = ws.cell(row=1, column=c)
                    mc.fill = GROUP_FILLS.get(key, HEADER_FILL)
                    mc.border = BORDER
            col += span
        start_row = 2

    # ---- header row ----
    for c, title in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    # ---- data rows ----
    for r, row in enumerate(rows, start=start_row + 1):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.alignment = WRAP_TOP
            cell.border = BORDER

    # ---- freeze panes below the header ----
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # ---- column widths ----
    for c, title in enumerate(headers, start=1):
        letter = ws.cell(row=start_row, column=c).column_letter
        if widths and c - 1 < len(widths):
            ws.column_dimensions[letter].width = widths[c - 1]
            continue
        longest = len(str(title))
        for row in rows:
            if c - 1 < len(row) and row[c - 1] is not None:
                longest = max(longest, len(str(row[c - 1])))
        ws.column_dimensions[letter].width = max(MIN_WIDTH, min(MAX_WIDTH, longest + 2))
